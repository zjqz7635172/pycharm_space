#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
# @Time    : 2021/6/29 10:25 上午
# @Author  : 玄凌
# @File    : excel_helper.py
# @Description : 功能描述
# @Software: PyCharm
"""

import os
import re
from os.path import isfile

from openpyxl import load_workbook, Workbook


def _get_excel_column_index(columnName: str) -> int:
    column_upper = columnName.upper()
    begin_char = 65
    column_chars = list(column_upper)
    column_index = 0
    length = len(column_chars)
    for i in range(length):
        char = column_chars[i]
        if char < 'A' or char > 'Z':
            return 0
        column_index = column_index + (ord(char) - begin_char) + i * 26
    return column_index


def _convert_value(value):
    """
    将单元格中数据，区分基本类型
    类似"true"/"false"(不区分大小写)转换为bool值
    长得像数字的转换为float类型
    其他（空格、空行）转换为None
    :param value: 单元格的值
    :return: 转换后的类型
    """
    value_str = str(value).lower()
    if value_str == 'true':
        return True
    elif value_str == 'false':
        return False
    elif re.match(r"^[+|-]?\d+.?\d*$", value_str):
        return float(value_str)
    elif re.match(r"^\s*$", value_str):
        return None
    else:
        return value


def _convert_value_v2(value):
    """
    将单元格中数据，不区分基本类型
    :param value: 单元格的值
    :return: str
    """
    return str(value)


class ExcelHelper:
    """
    Excel帮助类
    """

    @classmethod
    def _get_attr_index_dict(cls, clazz) -> dict:
        attrs = dir(clazz)
        attr_names = list(filter(lambda item: item[0] != '_', attrs))
        attr_dict = {}
        for attr_name in attr_names:
            attr_dict[_get_excel_column_index(attr_name)] = attr_name
        return attr_dict

    @classmethod
    def convert_2_class(cls, file_path, clazz, is_only_convert_str= False):
        """
        转换为class，可转换多张sheet表，转换为统一clazz对象
        过滤掉为空行
        :param file_path:Excel文件路径
        :param clazz:结果转换为clazz对象
        :return: 对象列表的列表，结构为[[clazz(),clazz()],[clazz()]]
        """
        if not file_path.endswith(".xlsx"):
            raise ValueError("文件必须为.xlsx结尾的Excel文件")
        if not isfile(file_path):
            raise FileNotFoundError("文件路径 {0} 不存在".format(file_path))
        work_book = load_workbook(file_path)
        result = []
        attr_dict = cls._get_attr_index_dict(clazz)
        for sheet_name in work_book.sheetnames:
            ws = work_book[sheet_name]

            # 获取表数据
            table = []
            for row in ws.iter_rows(min_row=2):
                # 实例化对象
                instance = clazz()
                for column_index in range(len(row)):
                    if column_index in attr_dict:
                        if is_only_convert_str is False:
                            setattr(instance, attr_dict[column_index], _convert_value(row[column_index].value))
                        else:
                            setattr(instance, attr_dict[column_index], _convert_value_v2(row[column_index].value))

                # 过滤空行（所有属性均为None的对象）
                is_valid = False
                for attr in instance.__dict__:
                    if not attr.startswith("_") and instance.__dict__[attr] is not None:
                        is_valid = True
                        break
                if is_valid:
                    table.append(instance)
            result.append(table)
        return result

    @classmethod
    def save(cls, file_path, tables, obj):
        if not file_path.endswith(".xlsx"):
            raise ValueError("文件必须为.xlsx结尾的Excel文件")
        work_book = Workbook()
        is_first = True
        for table in tables:
            if is_first:
                ws = work_book.active
                is_first = False
            else:
                ws = work_book.create_sheet()
            # 添加表头
            table_heads = []
            for attr in obj.__dict__:
                # 过滤"_"开头的属性
                if not attr.startswith("_"):
                    table_heads.append(attr)
            ws.append(table_heads)

            # 添加数据
            for row in table:
                data = []
                for head in table_heads:
                    data.append(getattr(row, head))
                ws.append(data)
        try:
            # 生成保存文件夹路径
            folder_index = max(file_path.rfind("\\"), file_path.rfind("/"))
            if folder_index != -1:
                folder_path = file_path[0:folder_index]
                if not os.path.exists(folder_path):
                    os.mkdir(folder_path)
            work_book.save(file_path)
        except Exception:
            raise OSError("创建Excel失败")


if __name__ == '__main__':
    pass
